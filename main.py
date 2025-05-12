from fastapi import FastAPI, UploadFile, WebSocket
import face_recognition
from fastapi.responses import JSONResponse
import numpy as np
import io
import os
from fastapi.middleware.cors import CORSMiddleware
import uuid
import pytesseract
from models.user_model import User
from openpyxl import load_workbook, Workbook
from src.id_scanner import IDCardScanner
from src.visitor_processor import extract_visitor_info, format_visitor_record
import cv2
from datetime import datetime
from controller.visitor_controller import VisitorController
from controller.id_card_controller import IdCardController
import re
import base64

upload_user_checkin_pics_directory= "known_users_images"
upload_user_checkin_encodings_directory= "known_users_encodings"
pytesseract.pytesseract.tesseract_cmd=r"C:\Program Files\Tesseract-OCR\tesseract.exe"
#model = torch.hub.load('ultralytics/yolov5', 'custom', path='cnic-yolov5.pt', force_reload=True)

WORD = re.compile(r"\w+")

visitor_controller= VisitorController()

# Define the file name
file_name = "visitors.xlsx"
file_path = os.path.join(os.getcwd(), file_name)

app = FastAPI()

# Allow CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],  # Frontend URL
    allow_credentials=True,
    allow_methods=["*"],  # or specify ['POST', 'GET', ...]
    allow_headers=["*"],  # or specify ['Authorization', 'Content-Type', ...]
)

# Get the current date and time in ISO 8601 format
current_iso_time = datetime.utcnow().isoformat() + "Z"  # Adding 'Z' to indicate UTC

@app.post('/verify/')
async def verify_user(file: UploadFile):
    try:
        image = await file.read()
        extension= file.filename.split('.')[1]
        v4= uuid.uuid4()
        # Load image and extract face encodings
        img = face_recognition.load_image_file(io.BytesIO(image))
        encodings = face_recognition.face_encodings(img)

        if not encodings:
            print("No face found in uploaded image")
            return JSONResponse(
                status_code=400,
                content={
                    "message": "No face detected. Please ensure your face is clearly visible, look directly at the camera, and try again in a well-lit area.",
                }
            )
        os.makedirs(upload_user_checkin_pics_directory,exist_ok=True)
        os.makedirs(upload_user_checkin_encodings_directory, exist_ok=True)

        verify_face= visitor_controller.verify_face(image=image, type="checkin")
        new_user_id= None
        checked_name= None
        if(verify_face["status"]==404):
            new_user_id= visitor_controller.create_user_face(image=image, encodings=encodings[0], extension=extension)
            # Save first face encoding (you can handle multiple faces if needed)
            encoding_file_path = os.path.join(upload_user_checkin_encodings_directory, f"{new_user_id}.npy")
            np.save(encoding_file_path, encodings[0])
            return JSONResponse(
                status_code=200,
                content={
                    "message":"Face not recognized. Your photo has been saved. Please complete the registration form to continue.",
                    "data":{
                        "new_user_id": str(new_user_id)
                    }
                }
            )
        else:
            current_time = datetime.utcnow().isoformat() + 'Z'
            final_message= None
            # Load Excel file
            wb = load_workbook(file_path)
            ws = wb.active  # Assumes single sheet
            for row in ws.iter_rows(min_row=2):  # Skip header
                if str(row[4].value) == verify_face["data"]["user_id"]:
                    if row[2].value is None: # add checkin if user comes for first time
                        row[2].value = current_time
                    else:
                        checked_name= row[0].value
                        #check checkedin and checkout array length
                        checkedin_value = row[2].value or ""
                        checkedout_value = row[3].value or ""

                        checkedin_length = len(checkedin_value.split(" | ")) if checkedin_value else 0
                        checkedout_length = len(checkedout_value.split(" | ")) if checkedout_value else 0

                        if checkedin_length == checkedout_length:
                            row[2].value = checkedin_value + (" | " if checkedin_value else "") + current_time  # checkin
                            final_message = "Welcome back, You have successfully checked in"
                        else:
                            row[3].value = checkedout_value + (" | " if checkedout_value else "") + current_time  # checkout
                            final_message = "Goodbye, You have successfully checked out."
                    break

            wb.save(filename=file_path)
            if checked_name is not None:
                    return JSONResponse(
                        status_code=200,
                        content={
                            "message": f'{final_message}, {checked_name}',
                        }
                    )
            else:
                return JSONResponse(
                    status_code=200,
                    content={
                        "message": f'No Face Found. Please Try again',
                    }
                )
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={
                "status": 500,
                "message": str(e)
            }
        )

@app.post('/scan-cnic/')
async def scan_cnic(file: UploadFile):
    try:
        # Read uploaded image
        contents = await file.read()
        nparr = np.frombuffer(contents, np.uint8)
        image = cv2.imdecode(nparr, cv2.COLOR_BGR2GRAY)
        
        # Initialize ID scanner
        scanner = IDCardScanner()
        
        # Process the image
        processed_image, ocr_result = scanner.process_image(image)
        
        # Extract text and confidence scores
        extracted_text = []
        for line in ocr_result:
            for word_info in line:
                text = word_info[1][0]
                confidence = word_info[1][1]
                extracted_text.append({
                    "text": text,
                    "confidence": float(confidence)
                })
        
        # Process visitor information
        ocr_response = {"extracted_text": extracted_text}
        name, cnic = extract_visitor_info(ocr_response)
        
        if not name or not cnic:
            return JSONResponse(
                status_code=400,
                content={
                    "message": "Could not extract required information from ID card",
                    "error": "Missing name or CNIC"
                }
            )
        
        # Format visitor record
        # visitor_record = format_visitor_record(name, cnic)
        
        # Save to Excel file
        # if os.path.exists(file_path):
        #     workbook = load_workbook(file_path)
        #     sheet = workbook.active
        # else:
        #     workbook = Workbook()
        #     sheet = workbook.active
        #     sheet.title = "Sheet1"
        #     headers = ["full_name", "cnic", "check_in", "check_out", "user_id"]
        #     sheet.append(headers)
        
        # # Append visitor data
        # sheet.append([
        #     visitor_record["full_name"],
        #     visitor_record["cnic"],
        #     visitor_record["check_in"],
        #     visitor_record["check_out"],
        #     visitor_record["user_id"]
        # ])
        
        # workbook.save(file_path)
        # v4= uuid.uuid4()
        
        return JSONResponse(
            status_code=200,
            content={
                "message": "ID card processed and visitor recorded successfully",
                "data": {
                    "full_name": name,
                    "cnic": cnic
                }
            }
        )
        
    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={
                "status": 500,
                "message": str(e)
            }
        )


@app.post('/create-user/')
async def create_user(user: User):
    try:
        # Check if the file exists
        if os.path.exists(file_path):
            workbook = load_workbook(file_path)
            sheet = workbook.active
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            # Write column headers only once when creating a new file
            headers = ["full_name", "cnic", "check_in", "check_out","user_id"]
            sheet.append(headers)

        # Append user data as a new row
        new_row = [user.full_name, user.cnic, user.check_in, "", user.user_id]
        sheet.append(new_row)

        # Save the updated Excel file
        workbook.save(file_path)

        return JSONResponse(
            status_code=200,
            content={"message": "User data saved successfully"}
        )

    except Exception as e:
        return JSONResponse(
            status_code=500,
            content={"message": str(e)}
        )
    
@app.websocket("/ws/detect-cnic")
async def websocket_endpoint(websocket: WebSocket):
    await websocket.accept()
    try:
        while True:
            data = await websocket.receive_text()
            image_bytes = base64.b64decode(data)
            nparr = np.frombuffer(image_bytes, np.uint8)
            img = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

            if img is None:
                await websocket.send_text("Invalid image")
                continue

            result = IdCardController().detect_id_card(img)
            await websocket.send_text(str(result))
    except Exception as e:
        print("WebSocket error:", e)
        await websocket.close()