import os
from openpyxl import Workbook, load_workbook
import uuid
import numpy as np
from numpy.typing import NDArray
import face_recognition
import io



upload_user_checkin_pics_directory= "known_users_images"
upload_user_checkin_encodings_directory= "known_users_encodings"

class VisitorController:

    #constructor
    def __init__(self):
        self.file_name = "visitors.xlsx"
        self.file_path = os.path.join(os.getcwd(), self.file_name)

        if os.path.exists(self.file_path):
            self.workbook = load_workbook(self.file_path)
            self.sheet = self.workbook.active
        else:
            self.workbook = Workbook()
            sheet = self.workbook.active
            self.sheet.title = "Sheet1"
            # Write column headers only once when creating a new file
            self.headers = ["full_name", "cnic", "check_in", "check_out","user_id"]
            self.sheet.append(self.headers)


    def update_visitor_checkin(self, user_id: str, check_in_time: str):
        # Iterate through rows to find the user by user_id
        for row in self.sheet.iter_rows(min_row=2, max_row=self.sheet.max_row, values_only=False):
            if row[4].value == user_id:  # Assuming user_id is in the 5th column
                row[2].value = check_in_time  # Update the check-in time (3rd column)
                break
        else:
            # If user_id is not found, raise an exception or handle as needed
            raise ValueError(f"User with user_id {user_id} not found.")

        # Save the updated Excel file
        self.workbook.save(self.file_path)

    def create_user_face(self, image: bytes, encodings: list[NDArray], extension: str):
        user_id= uuid.uuid4()
        file_location_for_known_users_images= f"./{upload_user_checkin_pics_directory}/{user_id}.{extension}"
        with open(file_location_for_known_users_images,"wb") as buffer:
            buffer.write(image)

        # Save first face encoding (you can handle multiple faces if needed)
        # encoding_file_path = os.path.join(upload_user_checkin_encodings_directory, f"{user_id}.npy")
        # np.save(encoding_file_path, encodings[0])

        return user_id
    
    def verify_face(self, image: bytes, type:str):
        # Load uploaded image and encode
        img = face_recognition.load_image_file(io.BytesIO(image))
        unknown_encodings = face_recognition.face_encodings(img)

        if not unknown_encodings:
            return {
                "status": 400,
                "message":"face can not be detected"
            }

        unknown_encoding = unknown_encodings[0]

        [known_encodings, known_ids]= self.get_all_encodings()

        # Compare the uploaded face with all known faces
        results = face_recognition.compare_faces(known_encodings, unknown_encoding, tolerance=0.5)

        for idx, match in enumerate(results):
            if match:
                matched_user_id = known_ids[idx]
                message= "Checked out"
                if("checkin" in type):
                    message= "Checked in"
                return {
                    "status": 200,
                    "message": message,
                    "data":{
                        "user_id": matched_user_id
                    }
                }
            
        return {
            "status": 404,
            "message": "User not verified",
        }
    

    def get_all_encodings(self):
        # Load all known encodings
        known_encodings = []
        known_ids = []

        for filename in os.listdir(upload_user_checkin_encodings_directory):
            if filename.endswith('.npy'):
                filepath = os.path.join(upload_user_checkin_encodings_directory, filename)
                encoding = np.load(filepath)
                known_encodings.append(encoding)
                user_id = filename.split('.')[0]  # uuid as user id
                known_ids.append(user_id)

        return [known_encodings,known_ids]
                